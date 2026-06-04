import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

class PointCalculator:
    """ポイント計算システム"""
    
    # 分数とポイントのマッピング
    DURATION_POINTS_MAP = {
        60: 1.0,
        75: 1.0,
        90: 1.5,
        120: 2.0,
        150: 2.5,
        180: 3.0,
        240: 4.0,
        300: 5.0,
        360: 6.0,
    }
    
    def __init__(self):
        self.duration_points_map = self.DURATION_POINTS_MAP
    
    def calculate_points(self, duration, has_designation=False):
        """
        分数からポイントを計算
        
        Args:
            duration (int): 勤務時間（分）
            has_designation (bool): 指名料の有無
        
        Returns:
            float: 計算されたポイント
        """
        points = self.duration_points_map.get(duration, 0)
        
        # 指名料がある場合は25%加算
        if has_designation:
            points = points * 1.25
        
        return round(points, 2)
    
    def get_duration_from_string(self, duration_str):
        """
        文字列から分数を抽出
        例: "60分", "1時間", "1:30" など
        
        Args:
            duration_str (str): 分数文字列
        
        Returns:
            int: 分数（整数）
        """
        if not duration_str:
            return None
        
        duration_str = str(duration_str).strip()
        
        # "分"で終わる場合
        if "分" in duration_str:
            try:
                return int(duration_str.replace("分", "").strip())
            except ValueError:
                return None
        
        # "時間"を含む場合
        if "時間" in duration_str:
            try:
                parts = duration_str.replace("時間", "").split(":")
                hours = int(parts[0])
                minutes = int(parts[1]) if len(parts) > 1 else 0
                return hours * 60 + minutes
            except (ValueError, IndexError):
                return None
        
        # "h" を含む場合（1.5h など）
        if "h" in duration_str.lower():
            try:
                hours = float(duration_str.replace("h", "").replace("H", "").strip())
                return int(hours * 60)
            except ValueError:
                return None
        
        # 数値のみの場合は分として解釈
        try:
            return int(float(duration_str))
        except ValueError:
            return None


class RewardExcelParser:
    """報酬明細Excelパーサー"""
    
    def __init__(self, excel_path, sheet_name=None):
        """
        Args:
            excel_path (str): Excelファイルのパス
            sheet_name (str): シート名（指定がない場合は最初のシート）
        """
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.wb = openpyxl.load_workbook(excel_path)
        
        if sheet_name:
            self.ws = self.wb[sheet_name]
        else:
            self.ws = self.wb.active
    
    def extract_data(self, name_col, duration_col, designation_col, header_row=1):
        """
        報酬明細から名前、分数、指名欄を抽出
        
        Args:
            name_col (int or str): 名前列（1 or 'A'）
            duration_col (int or str): 分数列
            designation_col (int or str): 指名欄列
            header_row (int): ヘッダー行番号
        
        Returns:
            list: [{'name': str, 'duration': int, 'has_designation': bool}, ...]
        """
        # 列文字に変換
        if isinstance(name_col, int):
            name_col = get_column_letter(name_col)
        if isinstance(duration_col, int):
            duration_col = get_column_letter(duration_col)
        if isinstance(designation_col, int):
            designation_col = get_column_letter(designation_col)
        
        data = []
        calculator = PointCalculator()
        
        for row in range(header_row + 1, self.ws.max_row + 1):
            name_cell = self.ws[f"{name_col}{row}"].value
            duration_cell = self.ws[f"{duration_col}{row}"].value
            designation_cell = self.ws[f"{designation_col}{row}"].value
            
            # 名前が空の場合はスキップ
            if not name_cell:
                continue
            
            # 分数を解析
            duration = calculator.get_duration_from_string(duration_cell)
            if duration is None:
                continue
            
            # 指名料の有無を判定
            has_designation = bool(designation_cell) and str(designation_cell).strip().upper() not in ["", "NO", "0", "FALSE"]
            
            data.append({
                'name': str(name_cell).strip(),
                'duration': duration,
                'has_designation': has_designation
            })
        
        return data
    
    def close(self):
        """ワークブックをクローズ"""
        self.wb.close()


class PointAggregationSheet:
    """ポイント集計表への自動入力"""
    
    def __init__(self, excel_path, sheet_name=None):
        """
        Args:
            excel_path (str): Excelファイルのパス
            sheet_name (str): シート名
        """
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.wb = openpyxl.load_workbook(excel_path)
        
        if sheet_name:
            self.ws = self.wb[sheet_name]
        else:
            self.ws = self.wb.active
    
    def insert_points(self, data_list, date, name_col_start=2, row_with_date_col=1):
        """
        ポイント集計表に日付・個人名でポイントを自動入力
        
        Args:
            data_list (list): [{'name': str, 'points': float}, ...] のリスト
            date (str or datetime): 日付
            name_col_start (int): 名前が開始される列番号
            row_with_date_col (int): 日付が記載される行番号
        
        Returns:
            bool: 成功したか
        """
        calculator = PointCalculator()
        
        # 日付の列を検索
        date_col = None
        date_str = str(date) if not isinstance(date, str) else date
        
        for col in range(1, self.ws.max_column + 1):
            cell_value = self.ws.cell(row=row_with_date_col, column=col).value
            if cell_value and str(cell_value).strip() == date_str:
                date_col = col
                break
        
        if date_col is None:
            print(f"警告: 日付'{date_str}'が見つかりません")
            return False
        
        # 各人のポイントを入力
        for item in data_list:
            name = item['name']
            points = item['points']
            
            # 名前の行を検索
            name_row = None
            for row in range(1, self.ws.max_row + 1):
                cell_value = self.ws.cell(row=row, column=name_col_start).value
                if cell_value and str(cell_value).strip() == name:
                    name_row = row
                    break
            
            if name_row is not None:
                # ポイントを入力
                self.ws.cell(row=name_row, column=date_col).value = points
                print(f"入力完了: {name} ({date}): {points}PT")
            else:
                print(f"警告: '{name}'が見つかりません")
        
        return True
    
    def save(self, output_path=None):
        """
        ワークブックを保存
        
        Args:
            output_path (str): 保存先パス（指定がない場合は元のファイルに上書き）
        """
        if output_path is None:
            output_path = self.excel_path
        
        self.wb.save(output_path)
        print(f"保存完了: {output_path}")
    
    def close(self):
        """ワークブックをクローズ"""
        self.wb.close()


def process_reward_to_points(reward_excel_path, aggregation_excel_path, 
                             date, reward_sheet=None, aggregation_sheet=None,
                             name_col=1, duration_col=2, designation_col=3,
                             name_col_agg=1, date_row=1):
    """
    報酬明細Excelからポイント集計表への全体処理
    
    Args:
        reward_excel_path (str): 報酬明細Excelファイルのパス
        aggregation_excel_path (str): ポイント集計表のパス
        date (str or datetime): 処理対象日付
        reward_sheet (str): 報酬明細のシート名
        aggregation_sheet (str): ポイント集計表のシート名
        name_col (int): 報酬明細の名前列
        duration_col (int): 報酬明細の分数列
        designation_col (int): 報酬明細の指名欄列
        name_col_agg (int): 集計表の名前列
        date_row (int): 集計表の日付行
    """
    calculator = PointCalculator()
    
    try:
        # ステップ1: 報酬明細から抽出
        print("=" * 50)
        print("ステップ1: 報酬明細から情報を抽出中...")
        print("=" * 50)
        
        parser = RewardExcelParser(reward_excel_path, reward_sheet)
        reward_data = parser.extract_data(name_col, duration_col, designation_col)
        parser.close()
        
        if not reward_data:
            print("エラー: 報酬明細からデータが抽出できません")
            return False
        
        print(f"抽出件数: {len(reward_data)}件\n")
        
        # ステップ2: ポイント計算
        print("=" * 50)
        print("ステップ2: ポイントを計算中...")
        print("=" * 50)
        
        point_data = []
        for item in reward_data:
            points = calculator.calculate_points(
                item['duration'], 
                item['has_designation']
            )
            designation_label = "指名料あり" if item['has_designation'] else "通常"
            print(f"{item['name']}: {item['duration']}分({designation_label}) → {points}PT")
            
            point_data.append({
                'name': item['name'],
                'points': points
            })
        
        print()
        
        # ステップ3: ポイント集計表に入力
        print("=" * 50)
        print("ステップ3: ポイント集計表に入力中...")
        print("=" * 50)
        
        aggregation = PointAggregationSheet(aggregation_excel_path, aggregation_sheet)
        success = aggregation.insert_points(point_data, date, name_col_agg, date_row)
        aggregation.save()
        aggregation.close()
        
        if success:
            print("\n✓ 処理完了しました！")
            return True
        else:
            print("\n✗ 処理中にエラーが発生しました")
            return False
    
    except Exception as e:
        print(f"\nエラーが発生しました: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    # 使用例
    # process_reward_to_points(
    #     reward_excel_path="報酬明細.xlsx",
    #     aggregation_excel_path="ポイント集計表.xlsx",
    #     date="2026-06-04",
    #     name_col=1,           # A列: 名前
    #     duration_col=2,       # B列: 分数
    #     designation_col=3     # C列: 指名欄
    # )
    pass
